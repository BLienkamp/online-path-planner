# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 14:43:37 2021

@author: benel
"""

# DO NOT EXECUTE THIS FILE JUST FOR FUN!!!!!
# THIS CODE RUNS THROUGH ALL SCENARIOS WITH THE TWO ONLINE ALGORITHMS AND THE BENCHMARK
# COMPLETELY EXECUTING THIS FILE TAKES MORE THAN 24 HOURS

import RunOA_no_info as no_info
import RunOA_with_info as with_info
import Gurobi as benchmark
from openpyxl import load_workbook

def run_tests(stop_assignments, layouts, scenarios, max_running_time, solution_file, algorithms_to_run, starting_from = 0):
    num = 1
    l = -1
    total_number_tests = len(stop_assignments) * len(layouts) * len(scenarios)
    number_algorithms = 0
    if algorithms_to_run[0]:
        number_algorithms += 1
    if algorithms_to_run[1]:
        number_algorithms += 1
    
    total_number_tests = total_number_tests * number_algorithms
    if algorithms_to_run[2]:
        total_number_tests += len(layouts) * len(scenarios)
    
    test_number = 0
    
    for layout in layouts:
        l +=1
        num = 1
        
        for scenario in scenarios:
            num += 1
            for stop_assignment in stop_assignments:

                print('#############################################')
                
                if algorithms_to_run[0]:
                    test_number += 1
                    print('\n')
                    print('(' + str(test_number) + '/' + str(total_number_tests) + ')')
                    # Run an instance with different parameters with no additional information
                    try:
                        run_no_info = no_info.runOA_no_info(stop_assignment, layout, scenario, show_timesteps= False)
                    #return obstacles, total_time, makespan, traffic_indicator, running_time, (walls_in_model, station_end, stops)
                    
                    except:
                        print('Error in layout ' + str(layout) + ' and scenario ' + str(scenario) + ' no info')
                        wb = load_workbook('Results/Results.xlsx')
                        ws = wb[stop_assignment]
                        
                        ws.cell(20*l+1,1,'Layout')
                        ws.cell(20*l+1,2,layout)
                        ws.cell(20*l+5,num, scenario)
                        ws.cell(20*l+6,1, 'Total time')
                        ws.cell(20*l+6,num, 'failed')
                        ws.cell(20*l+7,1, 'Makespan')
                        ws.cell(20*l+7,num, 'failed')
                        ws.cell(20*l+8,1, 'Traffic')
                        ws.cell(20*l+8,num, 'failed')
                        ws.cell(20*l+9,num, 'Running time')
                        ws.cell(20*l+9,num, 'failed')
                        #ws.cell(10,num, path)
                        wb.save('Results/Results.xlsx')
                        
                    else:
                        wb = load_workbook('Results/Results.xlsx')
                        
                        ws = wb[stop_assignment]
                        path = run_no_info[0]
                        total_time = run_no_info[1]
                        makespan = run_no_info[2]
                        traffic_indicator = run_no_info[3]
                        running_time = run_no_info[4]
                        
                        ws.cell(20*l+1,1,'Layout')
                        ws.cell(20*l+1,2,layout)
                        ws.cell(20*l+5,num, scenario)
                        ws.cell(20*l+6,1, 'Total time')
                        ws.cell(20*l+6,num, total_time)
                        ws.cell(20*l+7,1, 'Makespan')
                        ws.cell(20*l+7,num, makespan)
                        ws.cell(20*l+8,1, 'Traffic')
                        ws.cell(20*l+8,num, traffic_indicator)
                        ws.cell(20*l+9,num, 'Running time')
                        ws.cell(20*l+9,num, running_time)
                        #ws.cell(10,num, path)
                        wb.save('Results/Results.xlsx')
                
                if algorithms_to_run[1]:
                    test_number += 1
                    print('\n')
                    print('(' + str(test_number) + '/' + str(total_number_tests) + ')')
                    try:
                        run_with_info = with_info.runOA_with_info(stop_assignment, layout,scenario, perfect_estimation = False, show_timesteps = False)
                        
                    except:
                        print('Error in layout ' + str(layout) + ' and scenario ' + str(scenario) + ' with info')
                        wb = load_workbook('Results/Results.xlsx')
                        ws = wb[stop_assignment]
                        
                        ws.cell(20*l+13, num, scenario)
                        ws.cell(20*l+14,1, 'Total time')
                        ws.cell(20*l+14,num, 'failed')
                        ws.cell(20*l+15,1, 'Makespan')
                        ws.cell(20*l+15,num, 'failed')
                        ws.cell(20*l+16,1, 'Traffic')
                        ws.cell(20*l+16,num, 'failed')
                        ws.cell(20*l+17,1, 'Running time')
                        ws.cell(20*l+17,num, 'failed')
                        #ws.cell(10,num, path)
                        wb.save('Results/Results.xlsx')
                        
                    else:
                        wb = load_workbook('Results/Results.xlsx')
                        
                        ws = wb[stop_assignment]
                        path = run_with_info[0]
                        total_time = run_with_info[1]
                        makespan = run_with_info[2]
                        traffic_indicator = run_with_info[3]
                        running_time = run_with_info[4]
                        
                        ws.cell(20*l+13, num, scenario)
                        ws.cell(20*l+14,1, 'Total time')
                        ws.cell(20*l+14,num, total_time)
                        ws.cell(20*l+15,1, 'Makespan')
                        ws.cell(20*l+15,num, makespan)
                        ws.cell(20*l+16,1, 'Traffic')
                        ws.cell(20*l+16,num, traffic_indicator)
                        ws.cell(20*l+17,1, 'Running time')
                        ws.cell(20*l+17,num, running_time)
                        #ws.cell(10,num, path)
                        wb.save('Results/Results.xlsx')
                print('#############################################')
                print('\n')
                print('\n')
    
    l = -1
    if algorithms_to_run[2]:
        # We now run the benchmark
        for layout in layouts:
            l += 1
            num = 1 
            for scenario in scenarios:
                num += 1
                test_number += 1
                
                print('\n')
                print('(' + str(test_number) + '/' + str(total_number_tests) + ')')
                
                if test_number < starting_from:
                    continue
                wb = load_workbook('Results/Results.xlsx')
                
                # Find the smallest makespan we found so far for a scenario and use that to bound the max_time of the benchmark
                makespans_OA = []
                for stop_assignment in stop_assignments:
                    ws = wb[stop_assignment]
                    if isinstance(ws.cell(7, num).value, int):
                        makespans_OA.append(ws.cell(7, num).value)
                    if isinstance(ws.cell(15, num).value, int):
                        makespans_OA.append(ws.cell(15, num).value)
                
                # If no OA gets solved use 100
                if len(makespans_OA) < 1:
                    makespans_OA.append(100)

                try:
                    run_benchmark = benchmark.run_benchmark(layout, scenario, 'total time', max_run_time=20*60, max_time=min(makespans_OA))
                
                except:
                    print('Error in layout ' + str(layout) + ' and scenario ' + str(scenario) + ' benchmark')

                    ws = wb['Benchmark']
                    ws.cell(20*l+1,1,'Layout')
                    ws.cell(20*l+1,2,layout)
                    ws.cell(20*l+5,num, scenario)
                    ws.cell(20*l+6,1, 'Total time')
                    ws.cell(20*l+6,num, 'failed')
                    ws.cell(20*l+7,1, 'Makespan')
                    ws.cell(20*l+7,num, 'failed')
                    ws.cell(20*l+8,1, 'Traffic')
                    ws.cell(20*l+8,num, 'failed')
                    ws.cell(20*l+9,num, 'Running time')
                    ws.cell(20*l+9,num, 'failed')
                    #ws.cell(10,num, path)
                    wb.save('Results/Results.xlsx')
                    
                else:
                    path = run_benchmark[0]
                    total_time = run_benchmark[1]
                    makespan = run_benchmark[2]
                    traffic_indicator = run_benchmark[3]
                    running_time = run_benchmark[4]
                    
                    ws = wb['Benchmark']
                    ws.cell(20*l+1,1,'Layout')
                    ws.cell(20*l+1,2,layout)
                    ws.cell(20*l+5,num, scenario)
                    ws.cell(20*l+6,1, 'Total time')
                    ws.cell(20*l+6,num, total_time)
                    ws.cell(20*l+7,1, 'Makespan')
                    ws.cell(20*l+7,num, makespan)
                    ws.cell(20*l+8,1, 'Traffic')
                    ws.cell(20*l+8,num, traffic_indicator)
                    ws.cell(20*l+9,num, 'Running time')
                    ws.cell(20*l+9,num, running_time)
                    #ws.cell(10,num, path)
                    wb.save('Results/Results.xlsx')
                    
        
    return

# Uncomment this to execute the code
# run_tests(['ordered front to back', 'earliest front', 'random'],
#           ['1_9stops', '1_8stops', '1_7stops', '1_6stops', '1_5stops', '1_4stops',
#           '2_9stops', '2_8stops', '2_7stops', '2_6stops', '2_5stops', '2_4stops'],
#           ['1_rush', '2_rush', '3_rush', '4_rush', '5_rush', '6_rush', '7_rush', '8_rush', '9_rush', '10_rush', 
#            '11_rush', '12_rush', '13_rush', '14_rush', '15_rush', '16_rush', '17_rush', '18_rush', '19_rush', '20_rush',
#            '1_relax', '2_relax', '3_relax', '4_relax', '5_relax', '6_relax', '7_relax', '8_relax', '9_relax', '10_relax',
#            '11_relax', '12_relax', '13_relax', '14_relax', '15_relax', '16_relax', '17_relax', '18_relax', '19_relax', '20_relax'],0,0, 
#           (True, True, True), starting_from = 0)

